#!/usr/bin/env python3
"""
update_premiums.py  —  Refresh spot prices and option premiums in Selected Candidates.xlsx

Columns updated in a dated copy of sheet "CC stocks":
  D  Spot          — last regular-session close price
  F  Call Premium  — 35-45 delta call expiring in 4-11 days   (contract $ = price × 100)
  G  Put Premium   — 17-22 delta put expiring in 27-34 days    (contract $ = price × 100)

Stocks with no weekly option in the 4-11 day call window are skipped entirely.

Option price used: last traded → mark (bid/ask midpoint) → manual midpoint.
When no in-range delta exists, the premium is stored as an Excel interpolation formula:
  =w1*p1+w2*p2  (inverse-distance weighted from the two nearest deltas)

Usage
-----
  python update_premiums.py               # refresh all symbols
  python update_premiums.py NVDA AAPL     # refresh specific symbols only
"""

import sys
import os
from datetime import date, timedelta
from pathlib import Path

from dotenv import load_dotenv
import schwabdev
from openpyxl import load_workbook

# ── Configuration ─────────────────────────────────────────────────────────────
ENV_PATH     = Path(__file__).parent / ".env"
XLSX_PATH    = Path(r"C:\Users\adamr\Desktop\Selected Candidates.xlsx")
SOURCE_SHEET = "CC stocks"
DATA_START   = 7      # first data row (1-indexed, row 6 = header)
COL_SYMBOL   = 2      # B
COL_SPOT     = 4      # D
COL_CALL     = 6      # F
COL_PUT      = 7      # G

CALL_TARGET, CALL_LOW, CALL_HIGH = 40.0, 35.0, 45.0
PUT_TARGET,  PUT_LOW,  PUT_HIGH  = 20.0, 17.0, 22.0

# Expiration windows in days from today (inclusive)
CALL_MIN_DAYS, CALL_MAX_DAYS = 4, 11    # this/upcoming week
PUT_MIN_DAYS,  PUT_MAX_DAYS  = 27, 34   # ~4 weeks out

CONTRACT_SIZE = 100   # standard option contract = 100 shares

# Sentinel: no expiration found within the requested window
NO_EXPIRATION = object()
# ─────────────────────────────────────────────────────────────────────────────


def main() -> None:
    load_dotenv(ENV_PATH)
    symbols_filter = {s.upper() for s in sys.argv[1:]} or None

    today        = date.today()
    call_from    = today + timedelta(days=CALL_MIN_DAYS)
    call_to      = today + timedelta(days=CALL_MAX_DAYS)
    put_from     = today + timedelta(days=PUT_MIN_DAYS)
    put_to       = today + timedelta(days=PUT_MAX_DAYS)
    dated_name   = f"{SOURCE_SHEET} {today.month}.{today.day}.{str(today.year)[2:]}"

    _banner(today, call_from, call_to, put_from, put_to, dated_name, symbols_filter)

    # ── Load workbook and create dated sheet ──────────────────────────────────
    wb = load_workbook(XLSX_PATH)
    if dated_name in wb.sheetnames:
        del wb[dated_name]
    ws = wb.copy_worksheet(wb[SOURCE_SHEET])
    ws.title = dated_name

    # ── Schwabdev client (handles auth/token refresh automatically) ───────────
    client = schwabdev.Client(
        app_key=os.environ["SCHWAB_APP_KEY"],
        app_secret=os.environ["SCHWAB_APP_SECRET"],
        callback_url=os.environ.get("SCHWAB_CALLBACK_URL", "https://127.0.0.1"),
    )

    # ── Process rows ──────────────────────────────────────────────────────────
    ok = skipped_no_weekly = skipped_no_spot = 0

    for row in ws.iter_rows(min_row=DATA_START):
        sym_cell = row[COL_SYMBOL - 1]
        raw      = sym_cell.value
        if not raw or not isinstance(raw, str):
            continue
        symbol  = raw.strip().upper()
        row_num = sym_cell.row

        if symbols_filter and symbol not in symbols_filter:
            continue

        print(f"\n── {symbol} ──")

        # Call first — its expiration window determines whether the stock has weeklies
        call_val = _get_premium(client, symbol, call_from, call_to, "CALL")
        if call_val is NO_EXPIRATION:
            print(f"  ⊘  No weekly option in {CALL_MIN_DAYS}-{CALL_MAX_DAYS} day window — skipping symbol")
            skipped_no_weekly += 1
            continue

        # Spot
        spot = _get_spot(client, symbol)
        if spot is None:
            print(f"  ⚠  No spot price — skipping symbol")
            skipped_no_spot += 1
            continue

        # Put
        put_val = _get_premium(client, symbol, put_from, put_to, "PUT")
        if put_val is NO_EXPIRATION:
            put_val = None

        ws.cell(row=row_num, column=COL_SPOT).value = spot
        ws.cell(row=row_num, column=COL_CALL).value = call_val
        ws.cell(row=row_num, column=COL_PUT).value = put_val

        print(f"  Spot  : ${spot}")
        print(f"  Call  : {_fmt(call_val)}")
        print(f"  Put   : {_fmt(put_val)}")
        ok += 1

    wb.save(XLSX_PATH)
    print(f"\n{'─'*55}")
    print(f"Done — {ok} updated, {skipped_no_weekly} skipped (no weekly), "
          f"{skipped_no_spot} skipped (no spot).")
    print(f"Sheet '{dated_name}' saved to: {XLSX_PATH}")


# ── Market data ───────────────────────────────────────────────────────────────

def _get_spot(client: schwabdev.Client, symbol: str) -> float | None:
    try:
        resp = client.quote(symbol, fields="quote")
        resp.raise_for_status()
        q = resp.json().get(symbol, {}).get("quote", {})
        price = q.get("closePrice") or q.get("lastPrice")
        return round(float(price), 2) if price else None
    except Exception as e:
        print(f"  ✗ Quote error: {e}")
        return None


def _get_premium(
    client: schwabdev.Client,
    symbol: str,
    window_from: date,
    window_to: date,
    contract_type: str,
):
    """
    Fetch option chain and return premium as contract dollars (price × 100).

    Only expirations falling within [window_from, window_to] are considered.

    Returns:
      NO_EXPIRATION — no expiration exists in the window (e.g. no weekly option)
      float / str   — premium value or interpolation formula
      None          — expiration found but no usable contract data
    """
    try:
        resp = client.option_chains(
            symbol,
            contractType=contract_type,
            fromDate=window_from,
            toDate=window_to,
            includeUnderlyingQuote=False,
        )
        resp.raise_for_status()
        chain = resp.json()

        map_key = "callExpDateMap" if contract_type == "CALL" else "putExpDateMap"
        exp_map = chain.get(map_key, {})

        contracts = _expiration_in_window(exp_map, window_from, window_to)
        if contracts is None:
            return NO_EXPIRATION
        if not contracts:
            return None

        low, high, tgt = (CALL_LOW, CALL_HIGH, CALL_TARGET) if contract_type == "CALL" \
                    else (PUT_LOW,  PUT_HIGH,  PUT_TARGET)

        return _pick_premium(contracts, low, high, tgt)

    except Exception as e:
        print(f"  ✗ {contract_type} chain error: {e}")
        return None


def _expiration_in_window(exp_map: dict, window_from: date, window_to: date):
    """
    Return the contracts dict for the expiration inside [window_from, window_to]
    that sits closest to the window midpoint.

    Returns None if no expiration falls within the window at all.
    """
    midpoint = window_from + (window_to - window_from) / 2
    best_key, best_dist = None, float("inf")
    for key in exp_map:
        try:
            exp_dt = date.fromisoformat(key.split(":")[0])
        except ValueError:
            continue
        if exp_dt < window_from or exp_dt > window_to:
            continue
        dist = abs((exp_dt - midpoint).days)
        if dist < best_dist:
            best_dist, best_key = dist, key
    if best_key is None:
        return None
    return exp_map.get(best_key)


# ── Premium selection ─────────────────────────────────────────────────────────

def _pick_premium(
    contracts_by_strike: dict,
    low: float,
    high: float,
    target: float,
) -> float | str | None:
    """
    Scan all contracts in this expiration and select or interpolate a premium.

    Price priority: last traded → mark (bid/ask midpoint) → manual midpoint.
    Using 'last' as primary avoids the wide-spread distortion where mark can
    be far from any executable price (e.g. bid=$0.05, ask=$2.11 → mark=$1.08).

    Returns:
      float       — contract value (price/share × 100), rounded to 2dp, in-range
      str (Excel) — =w1*p1+w2*p2 interpolation formula
      None        — no usable contracts found
    """
    pairs: list[tuple[float, float]] = []  # (|delta| × 100, contract_value $)

    for _, contract_list in contracts_by_strike.items():
        for c in contract_list:
            raw_delta = c.get("delta")
            if raw_delta is None:
                continue
            delta = abs(raw_delta) * 100

            # Priority: last → mark → bid/ask midpoint
            price = c.get("last")
            if not price or price <= 0:
                price = c.get("mark")
            if not price or price <= 0:
                bid = c.get("bid") or 0
                ask = c.get("ask") or 0
                price = (bid + ask) / 2 if bid and ask else None
            if not price or price <= 0:
                continue

            pairs.append((delta, round(price * CONTRACT_SIZE, 2)))

    if not pairs:
        return None

    pairs.sort(key=lambda x: x[0], reverse=True)  # descending delta

    # ── In-range: pick closest to target ─────────────────────────────────────
    in_range = [(d, p) for d, p in pairs if low <= d <= high]
    if in_range:
        best = min(in_range, key=lambda x: abs(x[0] - target))
        return round(best[1], 2)

    # ── Out-of-range: interpolate from nearest on each side ──────────────────
    above = [(d, p) for d, p in pairs if d > high]
    below = [(d, p) for d, p in pairs if d < low]

    if above and below:
        c_above = min(above, key=lambda x: abs(x[0] - target))
        c_below = min(below, key=lambda x: abs(x[0] - target))
        return _interp_formula(c_above, c_below, target)

    # Only one side — use the single closest contract
    best = min(pairs, key=lambda x: abs(x[0] - target))
    return round(best[1], 2)


def _interp_formula(c1: tuple, c2: tuple, target: float) -> str:
    """
    Inverse-distance weighted interpolation between two contracts.
    The contract nearer to target receives the higher weight.

    Returns an Excel formula string: =w1*p1+w2*p2
    """
    d1, p1 = c1   # closer to target  (delta above high)
    d2, p2 = c2   # farther           (delta below low)
    dist1 = abs(d1 - target)
    dist2 = abs(d2 - target)
    total = dist1 + dist2

    w1 = round(dist2 / total, 4)  # nearer → higher weight
    w2 = round(1.0 - w1, 4)

    return f"={w1}*{p1}+{w2}*{p2}"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fmt(val) -> str:
    if val is None:
        return "⚠ no data"
    if isinstance(val, str):
        return f"[interpolated] {val}"
    return f"${val}"


def _banner(today, call_from, call_to, put_from, put_to, dated_name, symbols_filter) -> None:
    print("=" * 60)
    print("  update_premiums.py")
    print("=" * 60)
    print(f"  Today        : {today}")
    print(f"  Call expiry  : {call_from} → {call_to}  (35-45 delta)")
    print(f"  Put expiry   : {put_from} → {put_to}  (17-22 delta)")
    print(f"  Output sheet : '{dated_name}'")
    print(f"  Symbols      : {', '.join(sorted(symbols_filter)) if symbols_filter else 'all'}")
    print("  Note: stocks without a weekly option in the call window are skipped.")
    print("=" * 60)


if __name__ == "__main__":
    main()
